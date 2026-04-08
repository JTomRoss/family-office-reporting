from __future__ import annotations

import json
import re
from functools import lru_cache
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


_VISUAL_DICTIONARY_PATH = Path(__file__).resolve().parent / "asset_bucket_dictionary.json"
_EXCEL_DICTIONARY_PATH = (
    Path(__file__).resolve().parent
    / "Documentos"
    / "Excel"
    / "Diccionario de instrumentos.xlsx"
)
_SHORT_EXACT_ONLY_TOKENS = {"alt", "hy", "re", "rf", "rv"}
_COMPATIBILITY_ALIASES = {
    "SPDR": "RF IG Short",
    "ISHARES MSCI EM-ACC": "RV EM",
    "EMERGING MARKET EQUITIES": "RV EM",
    "EMERGING MARKETS EQUITIES": "RV EM",
    "NON-US EQUITY": "RV EM",
    "NON US EQUITY": "RV EM",
    "TOTAL NON-US EQUITY": "RV EM",
    "VAND USDCP1-3 USDA": "RF IG Short",
    "VANG USDCPBD USDA": "RF IG Long",
    "VANG USDCPBD USDA ACC": "RF IG Long",
}
_BUCKET_ALIASES = {
    "RF IG": "RF IG Short",
}


def _clean_excel_text(value) -> str | None:
    if value is None:
        return None
    text = str(value).replace("\xa0", " ").strip()
    return text or None


@lru_cache
def _load_visual_taxonomy() -> dict:
    return json.loads(_VISUAL_DICTIONARY_PATH.read_text(encoding="utf-8"))


@lru_cache
def _load_excel_dictionary_rows() -> tuple[tuple[str | None, ...], ...]:
    workbook = load_workbook(_EXCEL_DICTIONARY_PATH, data_only=True, read_only=True)
    try:
        sheet = workbook["diccionario inst"] if "diccionario inst" in workbook.sheetnames else workbook.active
        rows: list[tuple[str | None, ...]] = []
        for raw_row in sheet.iter_rows(values_only=True):
            row = tuple(_clean_excel_text(value) for value in raw_row)
            if any(row):
                rows.append(row)
        return tuple(rows)
    finally:
        workbook.close()


def _public_bucket_order() -> tuple[str, ...]:
    return tuple(str(bucket) for bucket in _load_visual_taxonomy().get("order", []))


def _is_public_bucket(value: str | None) -> bool:
    return bool(value) and value in set(_public_bucket_order())


def _canonical_bucket_name(value: str | None) -> str:
    raw = str(value or "").strip()
    if not raw:
        return ""
    alias = _BUCKET_ALIASES.get(raw.upper())
    if alias:
        return alias
    return raw


def _excel_row_bucket(row: tuple[str | None, ...]) -> str:
    canonical = row[0] if len(row) > 0 else None
    bucket_candidate = row[3] if len(row) > 3 else None
    if _is_public_bucket(canonical):
        return _canonical_bucket_name(canonical)
    if bucket_candidate:
        return _canonical_bucket_name(bucket_candidate)
    return _canonical_bucket_name(canonical)


def _looks_detail_label(value: str | None) -> bool:
    if not value:
        return False
    letters = [ch for ch in value if ch.isalpha()]
    if not letters:
        return True
    return not value.isupper()


def _excel_row_detail_label(row: tuple[str | None, ...], bucket: str) -> str:
    canonical = row[0] if len(row) > 0 else None
    second = row[1] if len(row) > 1 else None
    fifth = row[4] if len(row) > 4 else None
    if _is_public_bucket(canonical):
        if _looks_detail_label(fifth):
            return str(fifth)
        if _looks_detail_label(second):
            return str(second)
        return bucket
    if _looks_detail_label(fifth):
        return str(fifth)
    return bucket


def _iter_row_tokens(
    row: tuple[str | None, ...],
    *,
    bucket: str,
    detail_label: str,
) -> Iterable[str]:
    canonical = row[0] if len(row) > 0 else None
    if canonical:
        yield canonical
    if bucket:
        yield bucket
    if detail_label:
        yield detail_label

    if _is_public_bucket(canonical):
        extras = row[2:]
    else:
        extras = row[5:]

    for token in extras:
        if not token:
            continue
        lowered = token.lower()
        if lowered in {"rf", "rv"}:
            continue
        if lowered == "alt" and bucket != "Alternativos":
            continue
        yield token


def _unique_tokens(tokens: Iterable[str]) -> list[str]:
    unique: list[str] = []
    seen: set[str] = set()
    for token in tokens:
        cleaned = _clean_excel_text(token)
        if not cleaned:
            continue
        key = cleaned.casefold()
        if key in seen:
            continue
        seen.add(key)
        unique.append(cleaned)
    return unique


@lru_cache
def load_asset_taxonomy() -> dict:
    spec = dict(_load_visual_taxonomy())
    aliases: dict[str, str] = {}
    keyword_rules: list[dict[str, object]] = []
    detail_labels: dict[str, str] = {
        "PE": "PE",
        "RE": "RE",
    }

    for row in _load_excel_dictionary_rows():
        bucket = _excel_row_bucket(row)
        if not bucket:
            continue
        detail_label = _excel_row_detail_label(row, bucket)
        detail_labels[bucket] = detail_label

        row_tokens = _unique_tokens(_iter_row_tokens(row, bucket=bucket, detail_label=detail_label))
        for token in row_tokens:
            aliases.setdefault(token, bucket)

        keyword_tokens = [
            token
            for token in row_tokens
            if len(re.sub(r"\s+", "", token)) >= 4
            and token.casefold() not in _SHORT_EXACT_ONLY_TOKENS
        ]
        if keyword_tokens:
            keyword_rules.append({"bucket": bucket, "tokens": keyword_tokens})

    for alias, bucket in _COMPATIBILITY_ALIASES.items():
        aliases.setdefault(alias, bucket)

    spec["instrument_aliases"] = aliases
    spec["keyword_rules"] = keyword_rules
    spec["detail_labels"] = detail_labels
    return spec


def asset_bucket_order() -> list[str]:
    return list(load_asset_taxonomy().get("order", []))


def asset_bucket_colors() -> dict[str, str]:
    buckets = load_asset_taxonomy().get("buckets", {})
    return {
        str(bucket): str(meta.get("color") or "")
        for bucket, meta in buckets.items()
    }


def asset_bucket_color(bucket: str) -> str:
    return asset_bucket_colors().get(str(bucket), "")


def asset_bucket_detail_labels() -> dict[str, str]:
    return {
        str(bucket): str(label or bucket)
        for bucket, label in load_asset_taxonomy().get("detail_labels", {}).items()
    }


def asset_bucket_detail_label(bucket: str) -> str:
    normalized = str(bucket or "").strip()
    return asset_bucket_detail_labels().get(normalized, normalized)


def asset_bucket_series() -> list[tuple[str, str, str]]:
    return [
        (bucket, bucket, asset_bucket_color(bucket))
        for bucket in asset_bucket_order()
    ]


def coarse_asset_bucket_series() -> list[tuple[str, str, str]]:
    return [
        ("Cash, Deposits & Money Market", "Caja", asset_bucket_color("Caja")),
        ("Fixed Income", "Renta Fija", asset_bucket_color("RF IG Short")),
        ("Equities", "Renta Variable", asset_bucket_color("RV DM")),
    ]


def default_chart_color_sequence() -> list[str]:
    preferred_buckets = [
        "RV DM",
        "RF IG Short",
        "RV EM",
        "RF IG Long",
        "HY",
        "Non US RF",
        "Caja",
        "Alternativos",
        "Real Estate",
    ]
    return [
        color
        for color in (asset_bucket_color(bucket) for bucket in preferred_buckets)
        if color
    ]


def _classify_etf_asset_bucket_with_match(
    name: str,
    *,
    normalized_name: str | None = None,
) -> tuple[str, bool]:
    spec = load_asset_taxonomy()
    aliases = {
        str(key).strip().upper(): str(value).strip()
        for key, value in spec.get("instrument_aliases", {}).items()
        if str(key).strip() and str(value).strip()
    }
    compact_aliases = {
        re.sub(r"[^A-Z0-9]", "", key): value
        for key, value in aliases.items()
    }

    candidates = [str(normalized_name or "").strip(), str(name or "").strip()]
    for candidate in candidates:
        if not candidate:
            continue
        upper = candidate.upper()
        compact = re.sub(r"[^A-Z0-9]", "", upper)
        if upper in aliases:
            return aliases[upper], True
        if compact in compact_aliases:
            return compact_aliases[compact], True

    haystack = " ".join(value for value in candidates if value).lower()
    for rule in spec.get("keyword_rules", []):
        bucket = str(rule.get("bucket") or "").strip()
        tokens = [str(token).lower() for token in rule.get("tokens", []) if str(token).strip()]
        if bucket and any(token in haystack for token in tokens):
            return bucket, True

    return str(spec.get("default_bucket") or "RV DM"), False


def classify_etf_asset_bucket_with_match(
    name: str,
    *,
    normalized_name: str | None = None,
) -> tuple[str, bool]:
    return _classify_etf_asset_bucket_with_match(name, normalized_name=normalized_name)


def classify_etf_asset_bucket(name: str, *, normalized_name: str | None = None) -> str:
    bucket, _ = _classify_etf_asset_bucket_with_match(name, normalized_name=normalized_name)
    return bucket
